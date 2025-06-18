import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from base_datos import db
from utils.helpers import es_flotante
from threading import Timer
from openpyxl.utils import get_column_letter
import openpyxl
import datetime
import traceback

def formatear_precio(valor):
    if isinstance(valor, str):
        valor = valor.replace("$", "").replace(".", "").replace(",", ".")
    try:
        return "${:,.2f}".format(float(valor)).replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return valor

def limpiar_numero(valor):
    if isinstance(valor, str):
        valor = valor.replace("$", "").replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except Exception:
        return 0.0

def centrar_ventana(ventana):
    ventana.update_idletasks()
    width = ventana.winfo_width()
    height = ventana.winfo_height()
    x = (ventana.winfo_screenwidth() // 2) - (width // 2)
    y = (ventana.winfo_screenheight() // 2) - (height // 2)
    ventana.geometry(f"{width}x{height}+{x}+{y}")

def obtener_ultimo_cierre():
    try:
        with open("ultimo_cierre.txt", "r") as f:
            return f.read().strip()
    except FileNotFoundError:
        return "2000-01-01 00:00:00"

def iniciar_app():
    db.crear_tabla()
    ventana = tk.Tk()
    ventana.title("Control de Precios - Almacén")
    ventana.geometry("1100x600")
    ventana.minsize(900, 500)
    ventana.configure(bg="#f7f6f2")
    centrar_ventana(ventana)

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("TLabel", background="#f7f6f2", font=("Segoe UI", 12))
    style.configure("TButton", font=("Segoe UI", 11, "bold"), background="#4e9ca4", foreground="white")
    style.map("TButton", background=[("active", "#357376")])
    style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"), background="#4e9ca4", foreground="white")
    style.configure("Treeview", font=("Segoe UI", 11), rowheight=28, background="#ffffff", fieldbackground="#ffffff")

    main_frame = tk.Frame(ventana, bg="#f7f6f2")
    main_frame.pack(fill="both", expand=True)
    main_frame.grid_rowconfigure(0, weight=1)
    main_frame.grid_columnconfigure(1, weight=1)

    navbar = tk.Frame(main_frame, bg="#4e9ca4", width=180)
    navbar.grid(row=0, column=0, sticky="ns")
    navbar.grid_propagate(False)

    logo = tk.Label(navbar, text="🛒\nLO DE NAI", font=("Segoe UI", 16, "bold"), bg="#4e9ca4", fg="white", pady=20)
    logo.pack(pady=(30, 40))

    content = tk.Frame(main_frame, bg="#f7f6f2")
    content.grid(row=0, column=1, sticky="nsew")
    content.grid_rowconfigure(1, weight=1)
    content.grid_columnconfigure(0, weight=1)

    header_frame = tk.Frame(content, bg="#f7f6f2")
    header_frame.grid(row=0, column=0, sticky="ew", pady=(20, 10))
    header_frame.grid_columnconfigure(1, weight=1)

    titulo = tk.Label(header_frame, text="🛍️ Control de Precios", 
                    font=("Segoe UI", 20, "bold"), bg="#f7f6f2", fg="#4e9ca4")
    titulo.grid(row=0, column=0, sticky="w")

    search_frame = tk.Frame(header_frame, bg="#f7f6f2")
    search_frame.grid(row=0, column=1, sticky="e")

    tk.Label(search_frame, text="Buscar:", bg="#f7f6f2", 
            font=("Segoe UI", 10)).pack(side="left", padx=(0, 5))

    search_var = tk.StringVar()
    search_entry = tk.Entry(search_frame, font=("Segoe UI", 12), 
                        textvariable=search_var, width=25)
    search_entry.pack(side="left", ipady=2)

    search_btn = ttk.Button(search_frame, text="🔍", width=3,
                        command=lambda: buscar_producto(search_var.get()))
    search_btn.pack(side="left", padx=(5, 0))

    search_entry.bind("<Return>", lambda e: buscar_producto(search_var.get()))

    tabla_frame = tk.Frame(content, bg="#f7f6f2")
    tabla_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
    tabla_frame.grid_rowconfigure(0, weight=1)
    tabla_frame.grid_columnconfigure(0, weight=1)

    def mostrar_tabla():
        for widget in tabla_frame.winfo_children():
            widget.destroy()
        columnas = (
            "ID", "Descripción", "Precio Costo", "Porc. Utilidad", "Alicuota IVA",
            "Precio Venta", "Detalle Ext.", "Cod. Barras", "Cod. Producto"
        )
        tabla = ttk.Treeview(tabla_frame, columns=columnas, show="headings")
        for col in columnas:
            tabla.heading(col, text=col)
            tabla.column(col, anchor="center", width=130, minwidth=90, stretch=True)
        tabla.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=tabla.yview)
        tabla.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns")
        tabla_frame.grid_rowconfigure(0, weight=1)
        tabla_frame.grid_columnconfigure(0, weight=1)
        actualizar_datos_tabla(tabla)
        tabla_frame.tabla = tabla

    def actualizar_datos_tabla(tabla, productos=None):
        for row in tabla.get_children():
            tabla.delete(row)
        if productos is None:
            productos = db.obtener_productos()
        for prod in productos:
            tabla.insert("", "end", values=(
                prod[0],
                prod[1],
                formatear_precio(prod[2]),
                f"{float(prod[3]):.2f}%",
                f"{float(prod[4]):.2f}%",
                formatear_precio(prod[5]),
                prod[6] if prod[6] else "",
                prod[7] if prod[7] else "",
                prod[8] if prod[8] else ""
            ))

    def buscar_producto(termino_busqueda):
        termino = termino_busqueda.strip().lower()
        if not termino:
            actualizar_datos_tabla(tabla_frame.tabla)
            return
        try:
            productos = db.buscar_productos(termino)
            actualizar_datos_tabla(tabla_frame.tabla, productos)
        except Exception as e:
            messagebox.showerror("Error", f"Error al buscar: {str(e)}")
            print("Error completo:", traceback.format_exc())

    def mostrar_formulario_embebido(valores=None, producto_id=None):
        for widget in tabla_frame.winfo_children():
            widget.destroy()
        form_frame = tk.Frame(tabla_frame, bg="#f7f6f2")
        form_frame.grid(row=0, column=0, sticky="nsew", padx=60, pady=30)
        form_frame.grid_columnconfigure(1, weight=1)
        labels = [
            "Descripción", "Precio Costo", "Porcentaje Utilidad (%)", "Alicuota IVA",
            "Detalle Extendido (opcional)", "Código de Barras (opcional)", "Código de Producto (opcional)"
        ]
        keys = [
            "descripcion", "precio_costo", "porcentaje_utilidad", "alicuota_iva",
            "detalle_extendido", "codigo_barras", "codigo_producto"
        ]
        entries = {}
        for i, (label, key) in enumerate(zip(labels, keys)):
            tk.Label(form_frame, text=label+":", bg="#f7f6f2", fg="#333", font=("Segoe UI", 13, "bold" if i == 0 else "normal")).grid(row=i, column=0, sticky="e", padx=10, pady=8)
            if key == "alicuota_iva":
                iva_opciones = ["0%", "10,5%", "27%", "5%", "2,5%"]
                entries[key] = ttk.Combobox(form_frame, values=iva_opciones, font=("Segoe UI", 12), width=12, state="readonly")
                entries[key].set(valores[i] if valores else iva_opciones[0])
            else:
                entries[key] = tk.Entry(form_frame, font=("Segoe UI", 12), width=30)
                if valores:
                    entries[key].insert(0, valores[i])
            entries[key].grid(row=i, column=1, padx=10, pady=8, sticky="ew")
        tk.Label(form_frame, text="Precio Venta:", bg="#f7f6f2", fg="#333", font=("Segoe UI", 13, "bold")).grid(row=7, column=0, sticky="e", padx=10, pady=8)
        precio_venta_var = tk.StringVar()
        precio_venta_entry = tk.Entry(form_frame, font=("Segoe UI", 12, "bold"), width=15, textvariable=precio_venta_var, state="readonly", disabledforeground="#357376")
        precio_venta_entry.grid(row=7, column=1, padx=10, pady=8, sticky="w")
        if valores:
            precio_venta_var.set(valores[4] if len(valores) > 4 else "")
        def calcular_precio_venta(*args):
            try:
                costo = float(entries["precio_costo"].get())
                utilidad = float(entries["porcentaje_utilidad"].get())
                iva_str = entries["alicuota_iva"].get().replace("%", "").replace(",", ".")
                iva = float(iva_str) if iva_str else 0
                precio = costo + (costo * utilidad / 100)
                precio = precio + (precio * iva / 100)
                precio_venta_var.set(f"{precio:.2f}")
            except Exception:
                precio_venta_var.set("")
        entries["precio_costo"].bind("<KeyRelease>", calcular_precio_venta)
        entries["porcentaje_utilidad"].bind("<KeyRelease>", calcular_precio_venta)
        entries["alicuota_iva"].bind("<<ComboboxSelected>>", calcular_precio_venta)
        btn_frame = tk.Frame(form_frame, bg="#f7f6f2")
        btn_frame.grid(row=8, column=0, columnspan=2, pady=(20, 0))
        def volver_a_lista():
            mostrar_tabla()
        def cargar_producto():
            try:
                descripcion = entries["descripcion"].get().strip().title()
                precio_costo_str = entries["precio_costo"].get().replace(',', '.')
                porcentaje_utilidad_str = entries["porcentaje_utilidad"].get().replace(',', '.')
                alicuota_iva_str = entries["alicuota_iva"].get().replace('%', '').replace(',', '.')
                precio_venta_str = precio_venta_var.get().replace(',', '.')
                if not descripcion:
                    messagebox.showerror("Error", "La descripción es obligatoria", parent=form_frame)
                    return
                if not all(es_flotante(x) for x in [precio_costo_str, porcentaje_utilidad_str, alicuota_iva_str, precio_venta_str]):
                    messagebox.showerror("Error", "Los valores numéricos no son válidos", parent=form_frame)
                    return
                precio_costo = float(precio_costo_str)
                porcentaje_utilidad = float(porcentaje_utilidad_str)
                alicuota_iva = float(alicuota_iva_str)
                precio_venta = float(precio_venta_str)
                db.insertar_producto(
                    descripcion=descripcion,
                    precio_costo=precio_costo,
                    porcentaje_utilidad=porcentaje_utilidad,
                    alicuota_iva=alicuota_iva,
                    precio_venta=precio_venta,
                    detalle_extendido=entries["detalle_extendido"].get() or None,
                    codigo_barras=entries["codigo_barras"].get() or None,
                    codigo_producto=entries["codigo_producto"].get() or None
                )
                messagebox.showinfo("Éxito", "Producto cargado correctamente")
                mostrar_tabla()
            except Exception as e:
                messagebox.showerror("Error", f"Error al cargar producto: {str(e)}")
                print("Error completo:", traceback.format_exc())
        def guardar_edicion():
            try:
                descripcion = entries["descripcion"].get().strip().title()
                precio_costo_str = entries["precio_costo"].get().replace(',', '.')
                porcentaje_utilidad_str = entries["porcentaje_utilidad"].get().replace(',', '.')
                alicuota_iva_str = entries["alicuota_iva"].get().replace('%', '').replace(',', '.')
                precio_venta_str = precio_venta_var.get().replace(',', '.')
                if not descripcion:
                    messagebox.showerror("Error", "La descripción es obligatoria", parent=form_frame)
                    return
                if not all(es_flotante(x) for x in [precio_costo_str, porcentaje_utilidad_str, alicuota_iva_str, precio_venta_str]):
                    messagebox.showerror("Error", "Los valores numéricos no son válidos", parent=form_frame)
                    return
                precio_costo = float(precio_costo_str)
                porcentaje_utilidad = float(porcentaje_utilidad_str)
                alicuota_iva = float(alicuota_iva_str)
                precio_venta = float(precio_venta_str)
                db.editar_producto(
                    producto_id,
                    descripcion,
                    precio_costo,
                    porcentaje_utilidad,
                    alicuota_iva,
                    precio_venta,
                    entries["detalle_extendido"].get() or None,
                    entries["codigo_barras"].get() or None,
                    entries["codigo_producto"].get() or None
                )
                messagebox.showinfo("Éxito", "Producto editado correctamente")
                mostrar_tabla()
            except Exception as e:
                messagebox.showerror("Error", f"Error al editar producto: {str(e)}")
                print("Error completo:", traceback.format_exc())
        if producto_id:
            boton_volver = ttk.Button(btn_frame, text="Volver", command=volver_a_lista)
            boton_volver.pack(side="left", padx=10, ipadx=10, ipady=4)
            boton_guardar = ttk.Button(btn_frame, text="Guardar cambios", command=guardar_edicion)
            boton_guardar.pack(side="left", padx=10, ipadx=10, ipady=4)
        else:
            boton_volver = ttk.Button(btn_frame, text="Volver", command=volver_a_lista)
            boton_volver.pack(side="left", padx=10, ipadx=10, ipady=4)
            boton_cargar = ttk.Button(btn_frame, text="Cargar producto", command=cargar_producto)
            boton_cargar.pack(side="left", padx=10, ipadx=10, ipady=4)

    def mostrar_ventas_del_dia():
        ventas_win = tk.Toplevel()
        ventas_win.title("Ventas del día")
        ventas_win.geometry("700x400")
        ventas_win.configure(bg="#f7f6f2")
        centrar_ventana(ventas_win)

        ultimo_cierre = obtener_ultimo_cierre()
        ventas, detalles = db.obtener_ventas_desde(ultimo_cierre)

        cols = ("ID", "Fecha", "Cliente", "Pago", "Total")
        tabla_ventas = ttk.Treeview(ventas_win, columns=cols, show="headings")
        for col in cols:
            tabla_ventas.heading(col, text=col)
            tabla_ventas.column(col, width=120)
        tabla_ventas.pack(fill="both", expand=True, padx=10, pady=10)

        for v in ventas:
            tabla_ventas.insert("", "end", values=(v.id, v.fecha, v.cliente, v.pago, formatear_precio(v.total)))

        def ver_detalle(event):
            seleccionado = tabla_ventas.selection()
            if not seleccionado:
                return
            item = tabla_ventas.item(seleccionado[0])
            venta_id = item["values"][0]
            detalles_venta = [d for d in detalles if d.venta_id == venta_id]

            detalle_win = tk.Toplevel(ventas_win)
            detalle_win.title(f"Detalle de venta #{venta_id}")
            detalle_win.geometry("600x300")
            detalle_win.configure(bg="#f7f6f2")
            centrar_ventana(detalle_win)

            cols_det = ("Producto", "Precio Venta", "Cantidad", "Subtotal")
            tabla_det = ttk.Treeview(detalle_win, columns=cols_det, show="headings")
            for col in cols_det:
                tabla_det.heading(col, text=col)
                tabla_det.column(col, width=120)
            tabla_det.pack(fill="both", expand=True, padx=10, pady=10)

            for d in detalles_venta:
                tabla_det.insert("", "end", values=(d.descripcion, formatear_precio(d.precio_venta), d.cantidad, formatear_precio(d.subtotal)))

        tabla_ventas.bind("<Double-1>", ver_detalle)

    def mostrar_ventana_venta():
        venta_win = tk.Toplevel()
        venta_win.title("Nueva Venta")
        venta_win.geometry("1200x600")
        venta_win.configure(bg="#f7f6f2")
        centrar_ventana(venta_win)

        frame_buscar = tk.Frame(venta_win, bg="#f7f6f2")
        frame_buscar.pack(fill="x", pady=10)
        tk.Label(frame_buscar, text="Buscar producto:", bg="#f7f6f2").pack(side="left", padx=5)
        buscar_var = tk.StringVar()
        entry_buscar = tk.Entry(frame_buscar, textvariable=buscar_var, font=("Segoe UI", 12), width=30)
        entry_buscar.pack(side="left", padx=5)
        entry_buscar.focus()

        cols = ("ID", "Descripción", "Precio Venta", "Cod. Barras")
        tabla_busqueda = ttk.Treeview(frame_buscar, columns=cols, show="headings", height=3)
        for col in cols:
            tabla_busqueda.heading(col, text=col)
            tabla_busqueda.column(col, width=120)
        tabla_busqueda.pack(side="left", padx=5)

        def buscar_y_mostrar():
            termino = buscar_var.get().strip().lower()
            tabla_busqueda.delete(*tabla_busqueda.get_children())
            if not termino:
                return
            productos = db.buscar_productos(termino)
            for prod in productos:
                tabla_busqueda.insert("", "end", values=(prod[0], prod[1], formatear_precio(prod[5]), prod[7]))

        entry_buscar.bind("<KeyRelease>", lambda e: buscar_y_mostrar())

        btn_producto_nuevo = ttk.Button(frame_buscar, text="Producto nuevo", command=lambda: agregar_producto_nuevo())
        btn_producto_nuevo.pack(side="left", padx=10)

        frame_lista = tk.Frame(venta_win, bg="#f7f6f2")
        frame_lista.pack(fill="both", expand=True, pady=10)
        cols_venta = ("ID", "Descripción", "Precio Venta", "Cantidad", "Subtotal")
        tabla_venta = ttk.Treeview(frame_lista, columns=cols_venta, show="headings")
        for col in cols_venta:
            tabla_venta.heading(col, text=col)
            tabla_venta.column(col, width=110)
        tabla_venta.pack(fill="both", expand=True)

        def agregar_producto():
            seleccionado = tabla_busqueda.selection()
            if not seleccionado:
                return
            item = tabla_busqueda.item(seleccionado[0])
            prod_id, desc, precio_venta, cod_barras = item["values"]
            cantidad = 1
            for iid in tabla_venta.get_children():
                vals = tabla_venta.item(iid)["values"]
                if vals[0] == prod_id:
                    nueva_cant = vals[3] + 1 if isinstance(vals[3], int) else int(vals[3]) + 1
                    tabla_venta.item(iid, values=(
                        prod_id,
                        desc,
                        formatear_precio(limpiar_numero(precio_venta)),
                        nueva_cant,
                        formatear_precio(limpiar_numero(precio_venta) * nueva_cant)
                    ))
                    actualizar_total_local()
                    return
            tabla_venta.insert("", "end", values=(
                prod_id,
                desc,
                formatear_precio(limpiar_numero(precio_venta)),
                cantidad,
                formatear_precio(limpiar_numero(precio_venta) * cantidad)
            ))
            actualizar_total_local()

        tabla_busqueda.bind("<Double-1>", lambda e: agregar_producto())

        def agregar_producto_nuevo():
            win_nuevo = tk.Toplevel(venta_win)
            win_nuevo.title("Producto temporal")
            win_nuevo.geometry("350x180")
            win_nuevo.configure(bg="#f7f6f2")
            centrar_ventana(win_nuevo)

            tk.Label(win_nuevo, text="Nombre del producto:", bg="#f7f6f2").pack(pady=(15, 5))
            nombre_var = tk.StringVar()
            tk.Entry(win_nuevo, textvariable=nombre_var, font=("Segoe UI", 12)).pack(pady=5)

            tk.Label(win_nuevo, text="Precio de venta:", bg="#f7f6f2").pack(pady=(10, 5))
            precio_var = tk.StringVar()
            tk.Entry(win_nuevo, textvariable=precio_var, font=("Segoe UI", 12)).pack(pady=5)

            def agregar_a_venta():
                nombre = nombre_var.get().strip().title()
                precio = precio_var.get().replace(",", ".").strip()
                if not nombre or not precio:
                    messagebox.showerror("Error", "Debe ingresar nombre y precio.", parent=win_nuevo)
                    return
                try:
                    precio_float = float(precio)
                except ValueError:
                    messagebox.showerror("Error", "El precio debe ser un número.", parent=win_nuevo)
                    return
                cantidad = 1
                tabla_venta.insert("", "end", values=(
                    "TEMP",
                    nombre,
                    formatear_precio(precio_float),
                    cantidad,
                    formatear_precio(precio_float * cantidad)
                ))
                win_nuevo.destroy()
                actualizar_total_local()

            ttk.Button(win_nuevo, text="Agregar a venta", command=agregar_a_venta).pack(pady=15)

        frame_pago = tk.Frame(venta_win, bg="#f7f6f2")
        frame_pago.pack(fill="x", pady=10)
        tk.Label(frame_pago, text="Pago con:", bg="#f7f6f2").pack(side="left", padx=5)
        pago_var = tk.StringVar(value="Efectivo")
        ttk.Combobox(frame_pago, textvariable=pago_var, values=["Efectivo", "Transferencia"], state="readonly", width=15).pack(side="left", padx=5)
        tk.Label(frame_pago, text="Cliente:", bg="#f7f6f2").pack(side="left", padx=5)
        cliente_var = tk.StringVar()
        tk.Entry(frame_pago, textvariable=cliente_var, width=20).pack(side="left", padx=5)

        total_var = tk.StringVar(value="0,00")
        tk.Label(venta_win, text="Total: $", bg="#f7f6f2", font=("Segoe UI", 14, "bold")).pack(side="left", padx=(20,0))
        tk.Label(venta_win, textvariable=total_var, bg="#f7f6f2", font=("Segoe UI", 14, "bold")).pack(side="left")

        def actualizar_total_local():
            total = 0
            for iid in tabla_venta.get_children():
                vals = tabla_venta.item(iid)["values"]
                total += limpiar_numero(vals[4])
            total_var.set(formatear_precio(total))

        tabla_venta.bind("<<TreeviewSelect>>", lambda e: actualizar_total_local())
        tabla_venta.bind("<ButtonRelease-1>", lambda e: actualizar_total_local())

        def confirmar_venta():
            productos = []
            for iid in tabla_venta.get_children():
                vals = tabla_venta.item(iid)["values"]
                productos.append({
                    "id": vals[0],
                    "descripcion": vals[1],
                    "precio_venta": limpiar_numero(vals[2]),
                    "cantidad": vals[3],
                    "subtotal": limpiar_numero(vals[4])
                })
            if not productos:
                messagebox.showwarning("Venta", "No hay productos en la venta.")
                return
            pago = pago_var.get()
            cliente = cliente_var.get().strip().upper()
            total = limpiar_numero(total_var.get())
            fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db.registrar_venta(productos, pago, cliente, total, fecha)
            messagebox.showinfo("Venta", "Venta registrada correctamente.")
            venta_win.destroy()

        tk.Button(venta_win, text="Confirmar venta", command=confirmar_venta, bg="#4e9ca4", fg="white", font=("Segoe UI", 12, "bold")).pack(pady=10)

    def exportar_ventas_excel():
        ultimo_cierre = obtener_ultimo_cierre()
        ventas, detalles = db.obtener_ventas_desde(ultimo_cierre)
        if not ventas:
            messagebox.showinfo("Ventas", "No hay ventas para exportar desde el último cierre.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas del día"

        total_dia = sum(limpiar_numero(v.total) for v in ventas)
        ws.append(["TOTAL DEL DÍA", "", "", formatear_precio(total_dia)])
        ws.append(["Fecha", "Cliente", "Pago", "Total"])
        for v in ventas:
            ws.append([v.fecha, v.cliente, v.pago, formatear_precio(v.total)])

        ws2 = wb.create_sheet("Detalle")
        ws2.append(["Venta ID", "Producto", "Precio Venta", "Cantidad", "Subtotal"])
        for d in detalles:
            ws2.append([d.venta_id, d.descripcion, formatear_precio(d.precio_venta), d.cantidad, formatear_precio(d.subtotal)])

        for ws_ in [ws, ws2]:
            for col in ws_.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_.column_dimensions[column].width = max_length + 2

        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            wb.save(file)
            messagebox.showinfo("Exportar", "Ventas exportadas correctamente.")
            with open("ultimo_cierre.txt", "w") as f:
                f.write(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            messagebox.showinfo("Día finalizado", "El día ha sido finalizado. Las ventas del día ya no estarán visibles.")

    def mostrar_formulario_nav():
        mostrar_formulario_embebido()

    def eliminar_producto():
        tabla = tabla_frame.tabla
        seleccionado = tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Selecciona un producto para eliminar.")
            return
        item = tabla.item(seleccionado[0])
        id_producto = item["values"][0]
        descripcion = item["values"][1]
        if messagebox.askyesno("Eliminar", f"¿Seguro que deseas eliminar '{descripcion}'?"):
            db.eliminar_producto(id_producto)
            mostrar_tabla()

    def editar_producto():
        tabla = tabla_frame.tabla
        seleccionado = tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Selecciona un producto para editar.")
            return
        item = tabla.item(seleccionado[0])
        producto = item["values"]
        mostrar_formulario_embebido(valores=producto[1:], producto_id=producto[0])

    boton_nav_agregar = ttk.Button(navbar, text="Agregar producto", command=mostrar_formulario_nav)
    boton_nav_agregar.pack(fill="x", padx=20, pady=10, ipady=6)

    boton_nav_eliminar = ttk.Button(navbar, text="Eliminar producto", command=eliminar_producto)
    boton_nav_eliminar.pack(fill="x", padx=20, pady=10, ipady=6)

    boton_nav_editar = ttk.Button(navbar, text="Editar producto", command=editar_producto)
    boton_nav_editar.pack(fill="x", padx=20, pady=10, ipady=6)

    boton_nav_ver_ventas = ttk.Button(navbar, text="Ver ventas del día", command=mostrar_ventas_del_dia)
    boton_nav_ver_ventas.pack(fill="x", padx=20, pady=10, ipady=6)

    boton_nav_vender = ttk.Button(navbar, text="VENDER", command=mostrar_ventana_venta)
    boton_nav_vender.pack(fill="x", padx=20, pady=10, ipady=6)

    boton_nav_finalizar = ttk.Button(navbar, text="Finalizar día", command=exportar_ventas_excel)
    boton_nav_finalizar.pack(fill="x", padx=20, pady=10, ipady=6)

    search_timer = None
    def schedule_search(event=None):
        nonlocal search_timer
        if search_timer is not None:
            search_timer.cancel()
        search_timer = Timer(0.5, lambda: buscar_producto(search_var.get()))
        search_timer.start()
    search_entry.bind("<KeyRelease>", schedule_search)

    mostrar_tabla()
    ventana.mainloop()